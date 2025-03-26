import openpyxl
from datetime import datetime, timedelta

# Load the workbook and select the active sheet
file_path = "2025_reports_jan.xlsx"  # Update to your actual file path
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Arabic day names mapping
arabic_day_names = {
    "Saturday": "السبت",
    "Sunday": "الأحد",
    "Monday": "الاثنين",
    "Tuesday": "الثلاثاء",
    "Wednesday": "الأربعاء",
    "Thursday": "الخميس",
    "Friday": "الجمعة"
}

# Arabic month names mapping
arabic_month_names = {
    "January": "يناير",
    "February": "فبراير",
    "March": "مارس",
    "April": "أبريل",
    "May": "مايو",
    "June": "يونيو",
    "July": "يوليو",
    "August": "أغسطس",
    "September": "سبتمبر",
    "October": "أكتوبر",
    "November": "نوفمبر",
    "December": "ديسمبر"
}

# Set the month (March) and update control table (B2 for month, C2 for year)
month = 3  # March
merged_cell_range = sheet.merged_cells.ranges

# Access the top-left cell of the merged range
if "B2" in [str(cell) for cell in merged_cell_range]:
    # Find the merged range that contains B2
    for merge_range in merged_cell_range:
        if merge_range.start_cell.coordinate == "B2":
            # Access the top-left cell and update its value
            sheet[merge_range.start_cell.coordinate].value = month

# Access year value in C2 (if merged)
if "C2" in [str(cell) for cell in merged_cell_range]:
    for merge_range in merged_cell_range:
        if merge_range.start_cell.coordinate == "C2":
            base_year = sheet[merge_range.start_cell.coordinate].value
else:
    base_year = datetime.now().year

# Calculate the start date and number of days in the month
start_date = datetime(base_year, month, 1)
days_in_month = (datetime(base_year, month + 1, 1) - timedelta(days=1)).day

# Locate column headers for "ايام الاسبوع لشهر" (day names) and "التاريخ" (dates)
columns = {cell.value: cell.column for cell in sheet[1] if cell.value in ["ايام الاسبوع لشهر", "التاريخ"]}

if "ايام الاسبوع لشهر" not in columns or "التاريخ" not in columns:
    raise ValueError("Required columns 'ايام الاسبوع لشهر' or 'التاريخ' are missing in the sheet!")

# Populate the day names and dates
for day_offset in range(days_in_month):
    current_date = start_date + timedelta(days=day_offset)
    arabic_day_name = arabic_day_names[current_date.strftime('%A')]
    row = day_offset + 2  # Start populating from the second row

    # Update the day name and date columns
    # Update the month in cell C37
    sheet["C37"].value = month
    sheet.cell(row=row, column=columns["ايام الاسبوع لشهر"], value=arabic_day_name)
    sheet.cell(row=row, column=columns["التاريخ"], value=f"{current_date.month}/{current_date.day}")

# Save the updated file using Arabic month names
updated_file_path = f"{arabic_month_names[start_date.strftime('%B')]}_{base_year}.xlsx"
wb.save(updated_file_path)

print(f"Excel file updated successfully! Saved as {updated_file_path}")
