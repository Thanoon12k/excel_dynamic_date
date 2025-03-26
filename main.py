import openpyxl
from datetime import datetime, timedelta
# Arabic day names mapping
arabic_day_names = {
        "Saturday": "السبت",
        "Sunday": "الأحد",
        "Monday": "الاثنين",
        "Tuesday": "الثلاثاء",
        "Wednesday": "الأربعاء",
        "Thursday": "الخميس",
        "Friday": "الجمعة"
    }

# Arabic month names mapping
arabic_month_names = {
        1: "يناير",
        2: "فبراير",
        3: "مارس",
        4: "أبريل",
        5: "مايو",
        6: "يونيو",
        7: "يوليو",
        8: "أغسطس",
        9: "سبتمبر",
        10: "أكتوبر",
        11: "نوفمبر",
        12: "ديسمبر"
    }

def update_date_table_on_first_sheet(sheet,year, month):
    start_date = datetime(year, month, 1)
    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        current_date = start_date + timedelta(days=day - 1)
        sheet[f"B{41 + day - 1}"] = arabic_day_names[current_date.strftime('%A')]
        sheet[f"C{41 + day - 1}"] = f"{day}/{month}"

    print(f"table updated successfully for {arabic_month_names[month]} sheet {sheet.title}")
    return sheet

def generateMonthSheets(wb, year, month):
    delete_sheets = wb.sheetnames[1:]
    for sheet_name in delete_sheets:
        del wb[sheet_name]
    days_in_month = (datetime(year, month + 1, 1) - timedelta(days=1)).day
    for day in range(2, days_in_month + 1):
        new_sheet = wb.copy_worksheet(wb.worksheets[0])
        new_sheet.title = f"sheet_{day}"
    return wb


def change_sheet_dates_headers( year, month, wb):
    for i, sheet in enumerate(wb.worksheets):
        current_date = datetime(year, month, i + 1)
        day = i + 1
        current_date = datetime(year, month, day)
        sheet["F3"] = arabic_day_names[current_date.strftime('%A')]
        sheet["B37"] = year
        sheet["C37"] = month
        sheet["D37"] = day
        sheet.unmerge_cells("G3:I3")
        sheet["G3"] = current_date.strftime("%d/%m/%Y")
        sheet.merge_cells(f"G3:I3")
        sheet["G3"].alignment = openpyxl.styles.Alignment(horizontal="center")

def delete_date_table_from_other_sheets(wb):
    for sheet in wb.worksheets[1:]:
        for row in range(39, 76):
            sheet[f"B{row}"] = None
            sheet[f"C{row}"] = None
            sheet[f"B{row}"].fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            sheet[f"C{row}"].fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def createMonthReportFile(file_path, year, month):
    wb = openpyxl.load_workbook(file_path)
    wb.worksheets[0].title = f"sheet_1"
    update_date_table_on_first_sheet(wb.worksheets[0],year, month)

    wb=generateMonthSheets(wb, year, month)
    change_sheet_dates_headers( year, month, wb)
    delete_date_table_from_other_sheets(wb)
    output_file_path = f"{arabic_month_names[month]}_{year}.xlsx"
    wb.save(output_file_path)
    return output_file_path

if __name__ == "__main__":
    file_path = "original.xlsx"  # Update to your actual file path
    year = 2025
    month = 3  # March
    createMonthReportFile(file_path, year, month)
